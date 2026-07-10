"""
admin/exporter.py
------------------
Export helpers for the Admin Dashboard: CSV, Excel (.xlsx) and PDF.

CSV export uses only the standard library and is therefore always available.
XLSX and PDF export use ``openpyxl`` and ``reportlab`` respectively; both are
imported lazily and guarded, so if a deployment omits them the dashboard still
runs and simply reports that the format is unavailable (CSV remains offered).

Each function returns ``(bytes, mimetype, filename)`` ready to hand to Flask's
``send_file`` / ``Response``.
"""

from __future__ import annotations

import csv
import io
from datetime import datetime, timezone
from typing import Any, Dict, List, Sequence, Tuple

from utils.logging import logger

ExportResult = Tuple[bytes, str, str]


class ExportUnavailable(RuntimeError):
    """Raised when a requested export format's dependency is not installed."""


def _stamp() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S")


def _rows_to_matrix(
    rows: Sequence[Dict[str, Any]], columns: Sequence[str]
) -> List[List[Any]]:
    """Project a list of dicts onto an ordered column list."""
    return [[row.get(col, "") for col in columns] for row in rows]


def to_csv(rows: Sequence[Dict[str, Any]], columns: Sequence[str], name: str) -> ExportResult:
    """Export rows to CSV bytes."""
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(columns)
    writer.writerows(_rows_to_matrix(rows, columns))
    data = buffer.getvalue().encode("utf-8-sig")  # BOM for Excel friendliness
    return data, "text/csv", f"{name}-{_stamp()}.csv"


def to_xlsx(rows: Sequence[Dict[str, Any]], columns: Sequence[str], name: str) -> ExportResult:
    """Export rows to a styled .xlsx workbook (requires openpyxl)."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except Exception as exc:  # noqa: BLE001
        raise ExportUnavailable("openpyxl is not installed") from exc

    wb = Workbook()
    ws = wb.active
    ws.title = name[:31] or "Export"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="6D28D9")
    ws.append(list(columns))
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    for row in _rows_to_matrix(rows, columns):
        ws.append([_xlsx_safe(v) for v in row])

    for idx, col in enumerate(columns, start=1):
        width = max(len(str(col)), *(len(str(r.get(col, ""))) for r in rows)) if rows else len(str(col))
        ws.column_dimensions[get_column_letter(idx)].width = min(max(width + 2, 10), 60)
    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    return (
        buffer.getvalue(),
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        f"{name}-{_stamp()}.xlsx",
    )


def _xlsx_safe(value: Any) -> Any:
    """Coerce values openpyxl cannot store natively into strings."""
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def to_pdf(
    rows: Sequence[Dict[str, Any]], columns: Sequence[str], name: str, title: str = ""
) -> ExportResult:
    """Export rows to a landscape PDF table (requires reportlab)."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import cm
        from reportlab.platypus import (
            SimpleDocTemplate,
            Table,
            TableStyle,
            Paragraph,
            Spacer,
        )
    except Exception as exc:  # noqa: BLE001
        raise ExportUnavailable("reportlab is not installed") from exc

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=1 * cm,
        rightMargin=1 * cm,
        topMargin=1 * cm,
        bottomMargin=1 * cm,
    )
    styles = getSampleStyleSheet()
    elements: List[Any] = [
        Paragraph(title or name.replace("-", " ").title(), styles["Title"]),
        Paragraph(
            f"ME-HAAT Fashion AI Bot — generated {datetime.now(timezone.utc):%Y-%m-%d %H:%M UTC}",
            styles["Normal"],
        ),
        Spacer(1, 0.4 * cm),
    ]

    body = _rows_to_matrix(rows, columns)
    trimmed = [[_pdf_cell(v) for v in r] for r in body]
    table_data = [list(columns)] + (trimmed or [["No data"] + [""] * (len(columns) - 1)])
    table = Table(table_data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#6D28D9")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D1D5DB")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F3F4F6")]),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    elements.append(table)
    doc.build(elements)
    return buffer.getvalue(), "application/pdf", f"{name}-{_stamp()}.pdf"


def _pdf_cell(value: Any) -> str:
    """Stringify and length-cap a cell so wide fields don't overflow the page."""
    text = "" if value is None else str(value)
    return text if len(text) <= 90 else text[:87] + "…"


def export(
    fmt: str,
    rows: Sequence[Dict[str, Any]],
    columns: Sequence[str],
    name: str,
    title: str = "",
) -> ExportResult:
    """Dispatch to the requested export format ('csv' | 'xlsx' | 'pdf')."""
    fmt = (fmt or "csv").lower()
    if fmt == "csv":
        return to_csv(rows, columns, name)
    if fmt in {"xlsx", "excel"}:
        return to_xlsx(rows, columns, name)
    if fmt == "pdf":
        return to_pdf(rows, columns, name, title)
    logger.warning("ADMIN | Unknown export format %r; defaulting to CSV", fmt)
    return to_csv(rows, columns, name)
